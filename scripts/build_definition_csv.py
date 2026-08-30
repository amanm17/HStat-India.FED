#!/usr/bin/env python3
"""
Regenerate config/fed_sector_definition.csv from the FED Electronics
Sector Definition workbook.

The CSV — not this script, and not any API — is the single source of
truth for which HS codes HStat loads. This script exists only so that a
new edition of the workbook can be folded in without hand-editing 423
rows. Once the CSV exists it can be edited directly; the pipeline never
reads the workbook.

Usage:
    python scripts/build_definition_csv.py \
        --workbook "FED Electronics Sector Definition.xlsx" \
        --sheet "Sector Definition"
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit(
        "openpyxl is required: pip install openpyxl"
    )


ROOT = Path(__file__).resolve().parents[1]

DEST = ROOT / "config" / "fed_sector_definition.csv"


COLUMNS = [
    "hs6",
    "description",
    "product",
    "category",
    "segment",
    "dgcis_segment",
    "in_fed_definition",
    "in_old_fed_analysis",
    "in_telangana_study",
    "in_icea",
    "in_dgcis",
    "in_icrier",
    "world_exports_usd_bn",
    "share_of_world_trade",
    "comments",
    "search_terms",
]


SOURCE_HEADERS = {
    "hs6": ["HS Code"],
    "description": ["Description", "HS Code Description"],
    "product": ["Product"],
    "category": ["Category", "Product Category"],
    "segment": ["Segment"],
    "dgcis_segment": ["DGCIS Segments"],
    "in_fed_definition": [
        "Final FED Sector Definition",
        "Included in the new definition (Yes / No)",
    ],
    "in_old_fed_analysis": ["Included in old FED Analysis"],
    "in_telangana_study": ["Included in Telangana Study"],
    "in_icea": ["Included in ICEA"],
    "in_dgcis": ["Included in DGCIS"],
    "in_icrier": ["Included in ICRIER Report"],
    "comments": ["Comments", "Comment"],
}


def clean(value) -> str:
    if value is None:
        return ""

    text = re.sub(r"\s+", " ", str(value)).strip()

    return "" if text.lower() in {"none", "nan", "-"} else text


def title_case_category(value: str) -> str:
    """Category values arrive with inconsistent casing and trailing spaces."""
    if not value:
        return ""

    fixed = value.strip()

    return " ".join(
        word if word.isupper() else word.capitalize()
        for word in fixed.split()
    )


def yes_no(value: str) -> str:
    text = clean(value).lower()

    if text.startswith("y"):
        return "yes"

    if text.startswith("n"):
        return "no"

    return ""


def find_header_row(sheet, probe: str = "HS Code"):
    for row in sheet.iter_rows(min_row=1, max_row=12):
        values = [clean(cell.value) for cell in row]

        if probe in values:
            return row[0].row, values

    raise SystemExit(
        f"Could not find a '{probe}' header row in sheet {sheet.title!r}"
    )


def pick(headers: list[str], candidates: list[str]):
    for candidate in candidates:
        if candidate in headers:
            return headers.index(candidate)

    return None


def numeric(value) -> str:
    if value in (None, ""):
        return ""

    try:
        return f"{float(value):.6f}".rstrip("0").rstrip(".")
    except (TypeError, ValueError):
        return ""


def main():
    parser = argparse.ArgumentParser()

    parser.add_argument("--workbook", required=True)
    parser.add_argument("--sheet", default="Sector Definition")
    parser.add_argument("--out", default=str(DEST))

    args = parser.parse_args()

    book = openpyxl.load_workbook(args.workbook, data_only=True)

    if args.sheet not in book.sheetnames:
        raise SystemExit(
            f"Sheet {args.sheet!r} not in workbook. "
            f"Available: {book.sheetnames}"
        )

    sheet = book[args.sheet]

    header_row, headers = find_header_row(sheet)

    index = {
        field: pick(headers, candidates)
        for field, candidates in SOURCE_HEADERS.items()
    }

    if index["hs6"] is None:
        raise SystemExit("Workbook has no 'HS Code' column")

    value_column = next(
        (
            position
            for position, name in enumerate(headers)
            if name.lower().startswith("world exports")
        ),
        None,
    )

    share_column = next(
        (
            position
            for position, name in enumerate(headers)
            if name.lower().startswith("share of world")
        ),
        None,
    )

    seen: dict[str, dict] = {}
    duplicates: list[str] = []
    malformed: list[str] = []

    for row in sheet.iter_rows(
        min_row=header_row + 1,
        values_only=True,
    ):
        raw_code = clean(row[index["hs6"]]) if index["hs6"] < len(row) else ""

        if not raw_code:
            continue

        code = re.sub(r"\D", "", raw_code).zfill(6)

        if len(code) != 6:
            malformed.append(raw_code)
            continue

        def field(name: str) -> str:
            position = index.get(name)

            if position is None or position >= len(row):
                return ""

            return clean(row[position])

        record = {
            "hs6": code,
            "description": field("description"),
            "product": field("product"),
            "category": title_case_category(field("category")),
            "segment": field("segment"),
            "dgcis_segment": field("dgcis_segment"),
            "in_fed_definition": yes_no(field("in_fed_definition")),
            "in_old_fed_analysis": yes_no(field("in_old_fed_analysis")),
            "in_telangana_study": yes_no(field("in_telangana_study")),
            "in_icea": yes_no(field("in_icea")),
            "in_dgcis": yes_no(field("in_dgcis")),
            "in_icrier": yes_no(field("in_icrier")),
            "world_exports_usd_bn": (
                numeric(row[value_column])
                if value_column is not None and value_column < len(row)
                else ""
            ),
            "share_of_world_trade": (
                numeric(row[share_column])
                if share_column is not None and share_column < len(row)
                else ""
            ),
            "comments": field("comments"),
            "search_terms": "",
        }

        if code in seen:
            duplicates.append(code)
            continue

        seen[code] = record

    if malformed:
        print(
            f"Skipped {len(malformed)} malformed HS values: "
            + ", ".join(malformed[:8]),
            file=sys.stderr,
        )

    if duplicates:
        print(
            f"Skipped {len(duplicates)} duplicate HS codes: "
            + ", ".join(sorted(set(duplicates))[:8]),
            file=sys.stderr,
        )

    # Preserve any hand-written search_terms already in the CSV so a
    # workbook refresh never silently discards curated vocabulary.
    destination = Path(args.out)

    if destination.exists():
        with destination.open(newline="", encoding="utf-8") as handle:
            for existing in csv.DictReader(handle):
                code = clean(existing.get("hs6")).zfill(6)

                terms = clean(existing.get("search_terms"))

                if code in seen and terms:
                    seen[code]["search_terms"] = terms

    destination.parent.mkdir(parents=True, exist_ok=True)

    with destination.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=COLUMNS)

        writer.writeheader()

        for code in sorted(seen):
            writer.writerow(seen[code])

    included = sum(
        1
        for record in seen.values()
        if record["in_fed_definition"] == "yes"
    )

    print(f"Wrote {destination}")
    print(f"  HS-6 codes         : {len(seen)}")
    print(f"  In FED definition  : {included}")
    print(f"  Reference only     : {len(seen) - included}")


if __name__ == "__main__":
    main()
